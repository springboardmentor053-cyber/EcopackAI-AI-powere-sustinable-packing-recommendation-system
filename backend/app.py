from dotenv import load_dotenv
load_dotenv()

from flask import Flask, request, jsonify, render_template, send_file
from flask_cors import CORS

from pathlib import Path
import os
import io
from io import BytesIO
from datetime import datetime

import joblib
import pandas as pd

from backend.db import (
    load_materials,
    load_products,
    insert_recommendation_log,
    fetch_dashboard_summary,
    fetch_logs_df,
)

from src.feature_engineering import engineer_features

from dashboard.charts.bi_charts import (
    ensure_plot_dir,
    chart_top_materials,
    chart_avg_cost_by_category,
    chart_avg_co2_by_category,
)

# ---------- Excel (openpyxl) ----------
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table as XLTable, TableStyleInfo
from openpyxl.formatting.rule import ColorScaleRule

# ---------- PDF (reportlab) ----------
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import cm
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import (
    SimpleDocTemplate,
    Paragraph,
    Spacer,
    Table as PDFTable,
    TableStyle,
    PageBreak,
)
from datetime import datetime

BASE_DIR = Path(__file__).resolve().parent.parent  # project root
TEMPLATES_DIR = BASE_DIR / "frontend" / "templates"
STATIC_DIR = BASE_DIR / "frontend" / "static"


app = Flask(
    __name__,
    template_folder=str(TEMPLATES_DIR),
    static_folder=str(STATIC_DIR),
    static_url_path="/static",
)

CORS(app)

API_KEY = os.getenv("ECO_API_KEY", "dev-key-123")


def require_api_key():
    client_key = request.headers.get("X-API-KEY")
    if client_key != API_KEY:
        return jsonify({"error": "Unauthorized"}), 401
    return None, None

MODELS_DIR = BASE_DIR / "models"
RF_COST_PATH = MODELS_DIR / "rf_cost_model.pkl"
XGB_CO2_PATH = MODELS_DIR / "xgb_co2_model.pkl"
SCALER_PATH = MODELS_DIR / "scaler.pkl"


ML_FEATURES = [
    "product_weight_kg",
    "required_strength_score",
    "preferred_biodegradability_score",
    "strength_score",
    "weight_capacity_kg",
    "biodegradability_score",
    "recyclability_percent",
    "co2_emission_kg",
    "cost_per_unit_inr",
]


def load_artifacts():
    for p in [RF_COST_PATH, XGB_CO2_PATH, SCALER_PATH]:
        if not p.exists():
            raise FileNotFoundError(f"Missing file: {p}")

    rf_cost_model = joblib.load(RF_COST_PATH)
    xgb_co2_model = joblib.load(XGB_CO2_PATH)
    scaler_obj = joblib.load(SCALER_PATH)
    return rf_cost_model, xgb_co2_model, scaler_obj


rf_cost, xgb_co2, scaler = load_artifacts()


def _to_num(df: pd.DataFrame, cols: list[str]) -> pd.DataFrame:
    """Convert columns to numeric safely."""
    for c in cols:
        if c in df.columns:
            df[c] = pd.to_numeric(df[c], errors="coerce")
    return df


def _safe_float(x, default=0.0) -> float:
    try:
        v = float(x)
        if pd.isna(v):
            return float(default)
        return v
    except Exception:
        return float(default)


def _safe_int(x, default=10) -> int:
    try:
        v = int(float(x))
        return v
    except Exception:
        return int(default)


@app.get("/")
def landing():
    bi = fetch_dashboard_summary()
    return render_template("landing.html", bi=bi)


@app.get("/wizard")
def wizard_page():
    return render_template("wizard.html")


@app.get("/results")
def results_page():
    return render_template("results.html")


@app.get("/comparison")
def comparison_page():
    return render_template("comparison.html")



@app.get("/landing.html")
def landing_html():
    return render_template("landing.html")


@app.get("/wizard.html")
def wizard_html():
    return render_template("wizard.html")


@app.get("/results.html")
def results_html():
    return render_template("results.html")


@app.get("/comparison.html")
def comparison_html():
    return render_template("comparison.html")

@app.get("/dashboard")
def dashboard_page():
    data = fetch_dashboard_summary()

    # Generate charts into frontend/static/plots
    plot_dir = ensure_plot_dir(STATIC_DIR)

    try:
        if data.get("top_materials"):
            chart_top_materials(data["top_materials"], plot_dir / "top_materials.png")

        if data.get("by_category"):
            chart_avg_cost_by_category(data["by_category"], plot_dir / "avg_cost_by_category.png")
            chart_avg_co2_by_category(data["by_category"], plot_dir / "avg_co2_by_category.png")
    except Exception as e:
        print(f"[WARN] Chart generation failed: {e}")

    return render_template("dashboard.html", data=data)

@app.get("/api")
def api_root():
    return jsonify({"message": "EcoPackAI backend running ✅", "try": ["/api/health", "/api/recommend (POST)"]})



@app.get("/api/health")
def health():
    return jsonify({"status": "ok"})

@app.get("/export/excel")
def export_excel():
    # We need raw logs for the "Recent" + "Raw Logs" sheets
    data = fetch_dashboard_summary(include_raw=True)

    # --------- Helpers (local to this function) ----------
    THEME_DARK = "0B1E12"   # deep green
    THEME_GREEN = "1ED67B"  # accent green
    THEME_BORDER = "12321E" # dark border-ish
    WHITE = "FFFFFF"
    BLACK = "000000"

    def _safe_df(obj):
        df = pd.DataFrame(obj or [])
        return df

    def _set_col_widths(ws, max_width=45):
        # Nice looking fixed-ish widths (auto-fit but capped)
        for col_idx in range(1, ws.max_column + 1):
            col_letter = get_column_letter(col_idx)
            max_len = 0
            for row_idx in range(1, ws.max_row + 1):
                v = ws.cell(row=row_idx, column=col_idx).value
                s = "" if v is None else str(v)
                if len(s) > max_len:
                    max_len = len(s)
            ws.column_dimensions[col_letter].width = min(max_len + 3, max_width)

    def _style_header_row(ws, header_row=1):
        fill = PatternFill("solid", fgColor=THEME_GREEN)
        font = Font(bold=True, color=BLACK)
        align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for cell in ws[header_row]:
            cell.fill = fill
            cell.font = font
            cell.alignment = align

    def _add_filter(ws, start_row=1):
        if ws.max_row >= start_row and ws.max_column >= 1:
            ws.auto_filter.ref = f"A{start_row}:{get_column_letter(ws.max_column)}{ws.max_row}"

    def _freeze(ws, cell="A2"):
        ws.freeze_panes = cell

    def _add_excel_table(ws, table_name, start_row, start_col, end_row, end_col):
        # Excel table reference like A1:D20
        start = f"{get_column_letter(start_col)}{start_row}"
        end = f"{get_column_letter(end_col)}{end_row}"
        ref = f"{start}:{end}"

        tab = XLTable(displayName=table_name, ref=ref)
        
        style = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        tab.tableStyleInfo = style
        ws.add_table(tab)

    def _format_numbers(ws, currency_cols=None, decimal_cols=None, int_cols=None):
        currency_cols = set(currency_cols or [])
        decimal_cols = set(decimal_cols or [])
        int_cols = set(int_cols or [])

        for r in range(2, ws.max_row + 1):
            for c in range(1, ws.max_column + 1):
                cell = ws.cell(row=r, column=c)
                col_letter = get_column_letter(c)
                if cell.value is None:
                    continue
                if col_letter in currency_cols and isinstance(cell.value, (int, float)):
                    cell.number_format = '₹#,##0.00'
                if col_letter in decimal_cols and isinstance(cell.value, (int, float)):
                    cell.number_format = '0.00'
                if col_letter in int_cols and isinstance(cell.value, (int, float)):
                    cell.number_format = '0'

    # --------- Build DataFrames ----------
    total_requests = int(data.get("total_requests") or 0)
    avg_cost = float(data.get("avg_cost") or 0)
    avg_co2 = float(data.get("avg_co2") or 0)

    top_df = _safe_df(data.get("top_materials"))
    if not top_df.empty:
        # expected keys: recommended_material, cnt
        top_df = top_df.rename(columns={"recommended_material": "Material", "cnt": "Count"})

    cat_df = _safe_df(data.get("by_category"))
    if not cat_df.empty:
        # expected keys: product_category, avg_cost, avg_co2, cnt
        cat_df = cat_df.rename(
            columns={
                "product_category": "Category",
                "avg_cost": "Avg Cost (INR)",
                "avg_co2": "Avg CO₂ (kg)",
                "cnt": "Runs",
            }
        )

    raw_df = _safe_df(data.get("raw_logs"))
    # raw columns typically: id, created_at, product_category, product_weight_kg, fragility,
    # recommended_material, predicted_cost, predicted_co2
    # Keep it as-is but ensure consistent column order if present:
    preferred_cols = [
        "id", "created_at", "product_category", "product_weight_kg", "fragility",
        "recommended_material", "predicted_cost", "predicted_co2"
    ]
    if not raw_df.empty:
        cols_in = [c for c in preferred_cols if c in raw_df.columns]
        other_cols = [c for c in raw_df.columns if c not in cols_in]
        raw_df = raw_df[cols_in + other_cols]

    # Recent Recommendations (Snapshot) = last 25 rows
    recent_df = raw_df.copy()
    if not recent_df.empty and "created_at" in recent_df.columns:
        # Try sort by created_at desc
        try:
            recent_df = recent_df.sort_values("created_at", ascending=False)
        except Exception:
            pass
    if not recent_df.empty:
        recent_df = recent_df.head(25)

        # keep a compact snapshot view
        snapshot_cols_map = {
            "created_at": "Time",
            "product_category": "Category",
            "recommended_material": "Material",
            "predicted_cost": "Cost (INR)",
            "predicted_co2": "CO₂ (kg)",
        }
        snapshot_cols = [c for c in snapshot_cols_map.keys() if c in recent_df.columns]
        recent_df = recent_df[snapshot_cols].rename(columns=snapshot_cols_map)

    # --------- Write workbook ----------
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        wb = writer.book

        # ========== Sheet 1: Summary (Executive) ==========
        ws = wb.create_sheet("Summary", 0)

        # Title bar (merged)
        ws.merge_cells("A1:D2")
        title_cell = ws["A1"]
        title_cell.value = "EcoPackAI • Sustainability BI Report (Excel)"
        title_cell.fill = PatternFill("solid", fgColor=THEME_DARK)
        title_cell.font = Font(bold=True, color=WHITE, size=20)
        title_cell.alignment = Alignment(horizontal="left", vertical="center")

        # Section header
        ws.merge_cells("A4:D4")
        sec = ws["A4"]
        sec.value = "Executive Summary"
        sec.fill = PatternFill("solid", fgColor=THEME_DARK)
        sec.font = Font(bold=True, color=WHITE, size=12)
        sec.alignment = Alignment(horizontal="left", vertical="center")

        # Metric table header
        ws["A7"].value = "Metric"
        ws["B7"].value = "Value"
        ws.merge_cells("B7:D7")
        _style_header_row(ws, header_row=7)

        # Metrics rows
        rows = [
            ("Generated At", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
            ("Total Recommendations", total_requests),
            ("Avg Predicted Cost (INR)", avg_cost),
            ("Avg Predicted CO₂ (kg)", avg_co2),
        ]
        start_r = 8
        for i, (m, v) in enumerate(rows):
            r = start_r + i
            ws[f"A{r}"].value = m
            ws[f"B{r}"].value = v
            ws.merge_cells(f"B{r}:D{r}")

        # Format
        ws.column_dimensions["A"].width = 35
        ws.column_dimensions["B"].width = 25
        ws.column_dimensions["C"].width = 6
        ws.column_dimensions["D"].width = 6

        # Number formats
        ws["B10"].number_format = '₹#,##0.00'  # avg cost
        ws["B11"].number_format = '0.00'       # avg co2

        # Freeze so title+headers stay visible
        ws.freeze_panes = "A8"

        # ========== Sheet 2: Top Materials ==========
        ws2 = wb.create_sheet("Top Materials")
        if top_df.empty:
            ws2["A1"].value = "No data yet. Run recommendations to generate dashboard insights."
        else:
            top_df.to_excel(writer, sheet_name="Top Materials", index=False)
            ws2 = wb["Top Materials"]
            _style_header_row(ws2, 1)
            _freeze(ws2, "A2")
            _add_filter(ws2, 1)
            _set_col_widths(ws2)
            # Add Excel table
            _add_excel_table(ws2, "TopMaterialsTable", 1, 1, ws2.max_row, ws2.max_column)
            _format_numbers(ws2, int_cols=["B"])  # Count

        # ========== Sheet 3: Category Averages ==========
        ws3 = wb.create_sheet("Category Averages")
        if cat_df.empty:
            ws3["A1"].value = "No data yet. Run recommendations to generate dashboard insights."
        else:
            cat_df.to_excel(writer, sheet_name="Category Averages", index=False)
            ws3 = wb["Category Averages"]
            _style_header_row(ws3, 1)
            _freeze(ws3, "A2")
            _add_filter(ws3, 1)
            _set_col_widths(ws3)
            _add_excel_table(ws3, "CategoryAveragesTable", 1, 1, ws3.max_row, ws3.max_column)
            # Expected columns: A Category, B Avg Cost, C Avg CO2, D Runs
            _format_numbers(ws3, currency_cols=["B"], decimal_cols=["C"], int_cols=["D"])

        # ========== Sheet 4: Recent Recommendations ==========
        ws4 = wb.create_sheet("Recent Recommendations")
        if recent_df.empty:
            ws4["A1"].value = "No data yet. Run recommendations to generate recent snapshot."
        else:
            recent_df.to_excel(writer, sheet_name="Recent Recommendations", index=False)
            ws4 = wb["Recent Recommendations"]
            _style_header_row(ws4, 1)
            _freeze(ws4, "A2")
            _add_filter(ws4, 1)
            _set_col_widths(ws4)
            _add_excel_table(ws4, "RecentRecommendationsTable", 1, 1, ws4.max_row, ws4.max_column)
            # Time, Category, Material, Cost (INR), CO2 (kg)
            # Typically Cost in D, CO2 in E
            _format_numbers(ws4, currency_cols=["D"], decimal_cols=["E"])

        # ========== Sheet 5: Raw Logs (full) ==========
        ws5 = wb.create_sheet("Raw Logs")
        if raw_df.empty:
            ws5["A1"].value = "No raw logs available."
        else:
            raw_df.to_excel(writer, sheet_name="Raw Logs", index=False)
            ws5 = wb["Raw Logs"]
            _style_header_row(ws5, 1)
            _freeze(ws5, "A2")
            _add_filter(ws5, 1)
            _set_col_widths(ws5, max_width=55)
            _add_excel_table(ws5, "RawLogsTable", 1, 1, ws5.max_row, ws5.max_column)

            # Try to format known numeric columns if present
            # predicted_cost & predicted_co2 usually near end
            # Find their letters dynamically:
            headers = {ws5.cell(row=1, column=c).value: get_column_letter(c) for c in range(1, ws5.max_column + 1)}
            cost_col = headers.get("predicted_cost")
            co2_col = headers.get("predicted_co2")
            if cost_col:
                _format_numbers(ws5, currency_cols=[cost_col])
            if co2_col:
                _format_numbers(ws5, decimal_cols=[co2_col])

        # Remove default empty sheet if present (sometimes openpyxl creates one)
        if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
            wb.remove(wb["Sheet"])

    output.seek(0)
    return send_file(
        output,
        as_attachment=True,
        download_name="EcoPackAI_BI_Report.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

@app.get("/export/pdf")
def export_pdf():
    data = fetch_dashboard_summary(include_raw=True)

    # ---------- Theme colors (match your EcoPackAI vibe) ----------
    GREEN = colors.HexColor("#35FF8A")
    DARK = colors.HexColor("#07130B")
    PANEL = colors.HexColor("#0B1E12")
    PANEL2 = colors.HexColor("#0E2617")
    TEXT = colors.white
    MUTED = colors.Color(0.85, 0.92, 0.88)  # soft near-white

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        leftMargin=1.6 * cm,
        rightMargin=1.6 * cm,
        topMargin=1.4 * cm,
        bottomMargin=1.4 * cm,
        title="EcoPackAI Sustainability BI Report",
        author="EcoPackAI",
    )

    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(
        name="TitleEco",
        parent=styles["Title"],
        fontName="Helvetica-Bold",
        fontSize=18,
        textColor=TEXT,
        spaceAfter=8,
    ))
    styles.add(ParagraphStyle(
        name="SubEco",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=10,
        textColor=MUTED,
        spaceAfter=10,
    ))
    styles.add(ParagraphStyle(
        name="SectionEco",
        parent=styles["Heading2"],
        fontName="Helvetica-Bold",
        fontSize=12,
        textColor=TEXT,
        spaceBefore=10,
        spaceAfter=6,
    ))
    styles.add(ParagraphStyle(
        name="SmallMuted",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=9,
        textColor=MUTED,
        spaceAfter=6,
    ))

    def money(x):
        try:
            return f"{float(x):.2f}"
        except Exception:
            return "0.00"

    def num(x):
        try:
            return f"{float(x):.2f}"
        except Exception:
            return "0.00"

    # ---------- Header + KPIs ----------
    story = []

    story.append(Paragraph("EcoPackAI • Sustainability BI Report", styles["TitleEco"]))
    story.append(Paragraph(
        f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} • Data Source: PostgreSQL recommendation_logs",
        styles["SubEco"],
    ))
    story.append(Spacer(1, 6))

    total_requests = int(data.get("total_requests") or 0)
    avg_cost = money(data.get("avg_cost") or 0)
    avg_co2 = num(data.get("avg_co2") or 0)

    # KPI row as a styled table (3 “cards”)
    kpi_table = PDFTable(
        [
            [
                Paragraph("Total Recommendations", styles["SmallMuted"]),
                Paragraph("Avg Predicted Cost (INR)", styles["SmallMuted"]),
                Paragraph("Avg Predicted CO₂ (kg)", styles["SmallMuted"]),
            ],
            [
                Paragraph(f"<b>{total_requests}</b>", styles["TitleEco"]),
                Paragraph(f"<b>{avg_cost}</b>", styles["TitleEco"]),
                Paragraph(f"<b>{avg_co2}</b>", styles["TitleEco"]),
            ],
        ],
        colWidths=[(doc.width / 3.0)] * 3,
        hAlign="LEFT",
    )
    kpi_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), PANEL),
        ("BOX", (0, 0), (-1, -1), 1, GREEN),
        ("INNERGRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#1A3A2A")),
        ("TOPPADDING", (0, 0), (-1, -1), 10),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
        ("LEFTPADDING", (0, 0), (-1, -1), 12),
        ("RIGHTPADDING", (0, 0), (-1, -1), 12),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]))
    story.append(kpi_table)
    story.append(Spacer(1, 14))

    # ---------- Table helper ----------
    def styled_table(headers, rows, col_widths=None):
        tbl = PDFTable([headers] + rows, colWidths=col_widths, hAlign="LEFT")

        tbl.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), GREEN),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.black),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 9),
            ("ALIGN", (0, 0), (-1, 0), "LEFT"),
            ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
            ("TOPPADDING", (0, 0), (-1, 0), 6),

            ("BACKGROUND", (0, 1), (-1, -1), PANEL2),
            ("TEXTCOLOR", (0, 1), (-1, -1), colors.white),
            ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
            ("FONTSIZE", (0, 1), (-1, -1), 9),

            ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#1A3A2A")),
            ("BOX", (0, 0), (-1, -1), 1, colors.HexColor("#2EEB87")),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("LEFTPADDING", (0, 0), (-1, -1), 8),
            ("RIGHTPADDING", (0, 0), (-1, -1), 8),
            ("TOPPADDING", (0, 1), (-1, -1), 6),
            ("BOTTOMPADDING", (0, 1), (-1, -1), 6),
        ]))

        # Zebra striping (subtle)
        for i in range(1, len(rows) + 1):
            if i % 2 == 0:
                tbl.setStyle(TableStyle([
                    ("BACKGROUND", (0, i), (-1, i), colors.HexColor("#0B2416"))
                ]))

        return tbl

    # ---------- Top Materials ----------
    story.append(Paragraph("Top Recommended Materials", styles["SectionEco"]))
    top_materials = data.get("top_materials") or []
    if not top_materials:
        story.append(Paragraph("No recommendation history found yet.", styles["SmallMuted"]))
    else:
        rows = []
        for r in top_materials:
            rows.append([str(r.get("recommended_material", "")), str(r.get("cnt", 0))])

        tbl = styled_table(
            headers=["Material", "Count"],
            rows=rows,
            col_widths=[doc.width * 0.78, doc.width * 0.22],
        )
        story.append(tbl)

    story.append(Spacer(1, 14))

    # ---------- Category Averages ----------
    story.append(Paragraph("Category-wise Averages", styles["SectionEco"]))
    by_category = data.get("by_category") or []
    if not by_category:
        story.append(Paragraph("No category breakdown available.", styles["SmallMuted"]))
    else:
        rows = []
        for r in by_category:
            rows.append([
                str(r.get("product_category", "")),
                money(r.get("avg_cost", 0)),
                num(r.get("avg_co2", 0)),
                str(r.get("cnt", 0)),
            ])

        tbl = styled_table(
            headers=["Category", "Avg Cost (INR)", "Avg CO₂ (kg)", "Runs"],
            rows=rows,
            col_widths=[doc.width * 0.34, doc.width * 0.22, doc.width * 0.22, doc.width * 0.22],
        )
        story.append(tbl)

    story.append(Spacer(1, 14))

    # ---------- Recent Activity Snapshot (makes it executive-level) ----------
    story.append(Paragraph("Recent Recommendations (Snapshot)", styles["SectionEco"]))
    raw_logs = data.get("raw_logs") or []
    if not raw_logs:
        story.append(Paragraph("No recent activity found.", styles["SmallMuted"]))
    else:
        # Show latest 10
        rows = []
        for r in raw_logs[:10]:
            created = str(r.get("created_at", ""))[:19]
            rows.append([
                created,
                str(r.get("product_category", "")),
                str(r.get("recommended_material", ""))[:45],
                money(r.get("predicted_cost", 0)),
                num(r.get("predicted_co2", 0)),
            ])

        tbl = styled_table(
            headers=["Time", "Category", "Material", "Cost", "CO₂"],
            rows=rows,
            col_widths=[doc.width * 0.20, doc.width * 0.16, doc.width * 0.34, doc.width * 0.15, doc.width * 0.15],
        )
        story.append(tbl)

    # ---------- Footer / Page theme ----------
    def on_page(canvas_obj, doc_obj):
        canvas_obj.saveState()
        canvas_obj.setFillColor(DARK)
        canvas_obj.rect(0, 0, A4[0], A4[1], stroke=0, fill=1)

        # top accent line
        canvas_obj.setStrokeColor(GREEN)
        canvas_obj.setLineWidth(2)
        canvas_obj.line(doc.leftMargin, A4[1] - doc.topMargin + 6, A4[0] - doc.rightMargin, A4[1] - doc.topMargin + 6)

        # footer text
        canvas_obj.setFillColor(MUTED)
        canvas_obj.setFont("Helvetica", 9)
        canvas_obj.drawString(doc.leftMargin, 0.9 * cm, "EcoPackAI • Confidential • Generated by BI Export")
        canvas_obj.drawRightString(A4[0] - doc.rightMargin, 0.9 * cm, f"Page {doc_obj.page}")
        canvas_obj.restoreState()

    # Build PDF
    doc.build(story, onFirstPage=on_page, onLaterPages=on_page)

    buffer.seek(0)
    return send_file(
        buffer,
        as_attachment=True,
        download_name="EcoPackAI_BI_Report.pdf",
        mimetype="application/pdf",
    )
     
@app.post("/api/recommend")
def recommend():
    err, code = require_api_key()
    if err:
        return err, code

    data = request.get_json(silent=True) or {}
    product_name = str(data.get("product_name", "")).strip()
    if not product_name:
        return jsonify({"error": "product_name is required"}), 400

   
    top_n = _safe_int(data.get("top_n", 10), 10)
    top_n = max(1, min(top_n, 20))

   
    try:
        materials_df = load_materials()
        products_df = load_products()
    except Exception as e:
        return jsonify({"error": f"Database load failed: {str(e)}"}), 500

    
    products_df["product_name"] = products_df["product_name"].astype(str)
    materials_df["material_name"] = materials_df["material_name"].astype(str)

    
    materials_df = _to_num(
        materials_df,
        [
            "strength_score",
            "biodegradability_score",
            "cost_per_unit_inr",
            "weight_capacity_kg",
            "recyclability_percent",
            "co2_emission_kg",
        ],
    )
    products_df = _to_num(
        products_df,
        [
            "product_weight_kg",
            "required_strength_score",
            "preferred_biodegradability_score",
            "max_packaging_cost_inr",
        ],
    )

    match = products_df[products_df["product_name"].str.lower() == product_name.lower()]
    if match.empty:
        return jsonify({"error": f"Product '{product_name}' not found"}), 400

    selected_product = match.iloc[0]

    req_strength = _safe_float(selected_product.get("required_strength_score"), 0)
    pref_bio = _safe_float(selected_product.get("preferred_biodegradability_score"), 0)
    max_cost = _safe_float(selected_product.get("max_packaging_cost_inr"), 0)

    
    filtered_materials = materials_df.dropna(
        subset=["strength_score", "biodegradability_score", "cost_per_unit_inr"]
    ).copy()

    filtered_materials = filtered_materials[
        (filtered_materials["strength_score"] >= req_strength)
        & (filtered_materials["biodegradability_score"] >= pref_bio)
        & (filtered_materials["cost_per_unit_inr"] <= max_cost)
    ].copy()

    if filtered_materials.empty:
        return jsonify({"error": "No materials matched your requirements"}), 200

   
    fe_df = engineer_features(filtered_materials, selected_product)
    if "material_suitability_score" not in fe_df.columns:
        return jsonify({"error": "Feature engineering failed: 'material_suitability_score' missing"}), 500

    
    ranked = fe_df.sort_values("material_suitability_score", ascending=False).head(top_n).copy()

    ranked["product_weight_kg"] = _safe_float(selected_product.get("product_weight_kg"), 0)
    ranked["required_strength_score"] = req_strength
    ranked["preferred_biodegradability_score"] = pref_bio

    
    for c in ML_FEATURES:
        if c not in ranked.columns:
            ranked[c] = 0.0

    ranked = _to_num(ranked, ML_FEATURES).fillna(0)

    
    X_pred = ranked[ML_FEATURES].copy()

    ranked["pred_cost_inr"] = rf_cost.predict(X_pred)
    ranked["pred_co2_kg"] = xgb_co2.predict(scaler.transform(X_pred))


    max_co2 = float(pd.to_numeric(ranked["pred_co2_kg"], errors="coerce").max() or 1.0)
    if max_co2 <= 0:
        max_co2 = 1.0

    ranked["co2_score_norm"] = 1 - (ranked["pred_co2_kg"] / max_co2)
    ranked["environment_score"] = (
        0.40 * (ranked["biodegradability_score"] / 10)
        + 0.40 * (ranked["recyclability_percent"] / 100)
        + 0.20 * ranked["co2_score_norm"]
    ) * 100

 
    results = []
    for i, row in ranked.reset_index(drop=True).iterrows():
        results.append(
            {
                "rank": int(i + 1),
                "material_name": str(row.get("material_name", "")),
                "pred_cost_inr": float(row.get("pred_cost_inr", 0.0)),
                "pred_co2_kg": float(row.get("pred_co2_kg", 0.0)),
                "recyclability_percent": float(row.get("recyclability_percent", 0.0)),
                "biodegradability_score": float(row.get("biodegradability_score", 0.0)),
                "suitability_score": float(row.get("material_suitability_score", 0.0)),
                "environment_score": float(row.get("environment_score", 0.0)),
            }
        )
        # -----------------------------
    # Module 7: Log top recommendation for BI Dashboard
    # -----------------------------
    try:
        if results:
            top1 = results[0]  # rank 1
            insert_recommendation_log(
                product_category=str(selected_product.get("product_category", "")),
                product_weight_kg=_safe_float(selected_product.get("product_weight_kg"), 0),
                fragility=str(selected_product.get("fragility_level", "")),
                recommended_material=str(top1.get("material_name", "")),
                predicted_cost=float(top1.get("pred_cost_inr", 0.0)),
                predicted_co2=float(top1.get("pred_co2_kg", 0.0)),
            )
    except Exception as e:
        # Don't break recommendations if logging fails
        print(f"[WARN] BI logging failed: {e}")
    # ✅ Log ALL recommendations (top_n) to DB (not just top-1)
    try:
        for rec in results:
            insert_recommendation_log(
                product_category=str(selected_product.get("product_category", "")),
                product_weight_kg=_safe_float(selected_product.get("product_weight_kg"), 0),
                fragility=str(selected_product.get("fragility_level", "")),
                recommended_material=str(rec.get("material_name", "")),
                predicted_cost=_safe_float(rec.get("pred_cost_inr", 0)),
                predicted_co2=_safe_float(rec.get("pred_co2_kg", 0)),
            )
    except Exception as e:
        print(f"[WARN] Failed to log recommendations: {e}")



    return jsonify(
        {
            "product": {
                "product_name": str(selected_product.get("product_name", "")),
                "product_category": str(selected_product.get("product_category", "")),
                "product_weight_kg": _safe_float(selected_product.get("product_weight_kg"), 0),
                "fragility_level": str(selected_product.get("fragility_level", "")),
                "temperature_sensitive": str(selected_product.get("temperature_sensitive", "")),
            },
            "top_n": top_n,
            "recommendations": results,
        }
    )


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=True)
