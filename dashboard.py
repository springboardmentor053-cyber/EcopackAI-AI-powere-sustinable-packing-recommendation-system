"""
Business Intelligence Dashboard Module
Provides analytics for CO2 reduction, cost savings, and material usage trends
"""

import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import plotly.graph_objects as go
import plotly.express as px
from plotly.subplots import make_subplots
from datetime import datetime, timedelta
import json
import os
from io import BytesIO

# Try to import optional dependencies for export
try:
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image, PageBreak
    from reportlab.lib.units import inch
    HAS_REPORTLAB = True
except ImportError:
    HAS_REPORTLAB = False

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils.dataframe import dataframe_to_rows
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


class EcopackDashboard:
    """Analytics dashboard for sustainability metrics"""
    
    def __init__(self, product_csv='c:/EcopackAI/Dataset/product.csv', 
                 material_csv='c:/EcopackAI/Dataset/material.csv'):
        """Initialize dashboard with data from CSV files"""
        try:
            self.products_df = pd.read_csv(product_csv)
            self.materials_df = pd.read_csv(material_csv)
            self.data_loaded = True
        except FileNotFoundError:
            print("Warning: CSV files not found. Using sample data.")
            self.generate_sample_data()
            self.data_loaded = False
    
    def generate_sample_data(self):
        """Generate realistic sample data for dashboard"""
        self.products_df = pd.DataFrame({
            'product_id': range(1, 21),
            'product_name': [
                'Smartphone box', 'Laptop shipping pack', 'Earbuds box',
                'Electronics shipper', 'Snack wrapper', 'Chips bag',
                'Ready-meal tray', 'Beverage can', 'Jam jar', 'Coffee bag',
                'T-shirt box', 'Shoe box', 'Cosmetics box', 'Medicine bottle',
                'Food container', 'Protective case', 'Tissue box', 'Toy box',
                'Book mailer', 'Jewelry box'
            ],
            'sector': ['Electronics']*4 + ['Food']*3 + ['Food']*2 + ['General']*11,
            'co2_emission_score': np.random.randint(50, 95, 20),
            'biodegradability_score': np.random.randint(20, 95, 20),
            'recyclability_percent': np.random.randint(20, 95, 20),
            'cost_score': np.random.randint(2, 5, 20),
            'reuse_potential_score': np.random.randint(1, 5, 20)
        })
        
        self.materials_df = pd.DataFrame({
            'material_id': range(1, 16),
            'material_name': [
                'Recycled cardboard', 'Molded pulp', 'Kraft paper',
                'Starch-based foam', 'Mycelium composite', 'Bagasse fiber',
                'PLA bioplastic', 'Compostable bioplastic', 'Seaweed film',
                'Recycled plastic', 'Aluminum', 'Glass', 'Paper padding',
                'Biodegradable wrap', 'Cork padding'
            ],
            'material_type': ['Fiber-based']*3 + ['Bio-based']*6 + ['Other']*6,
            'co2_emission_score': np.random.randint(50, 90, 15),
            'biodegradability_score': np.random.randint(20, 95, 15),
            'recyclability_percent': np.random.randint(20, 85, 15)
        })
    
    def get_co2_metrics(self):
        """Calculate CO2 reduction metrics"""
        avg_co2_product = self.products_df['co2_emission_score'].mean()
        avg_co2_material = self.materials_df['co2_emission_score'].mean()
        
        best_products = self.products_df.nsmallest(5, 'co2_emission_score')[
            ['product_name', 'co2_emission_score']
        ].to_dict('records')
        
        best_materials = self.materials_df.nsmallest(5, 'co2_emission_score')[
            ['material_name', 'co2_emission_score']
        ].to_dict('records')
        
        reduction_potential = 30
        
        return {
            'avg_product_score': round(avg_co2_product, 2),
            'avg_material_score': round(avg_co2_material, 2),
            'reduction_potential_percent': reduction_potential,
            'best_products': best_products,
            'best_materials': best_materials,
            'total_emissions_baseline': round(avg_co2_product * len(self.products_df), 2),
            'total_emissions_optimized': round(
                avg_co2_product * len(self.products_df) * (1 - reduction_potential/100), 2
            )
        }
    
    def get_cost_metrics(self):
        """Calculate cost savings metrics"""
        avg_cost = self.products_df['cost_score'].mean()
        cost_distribution = self.products_df['cost_score'].value_counts().to_dict()
        
        potential_savings = 25
        total_cost_baseline = self.products_df['cost_score'].sum() * 100
        total_cost_optimized = total_cost_baseline * (1 - potential_savings/100)
        
        sector_costs = self.products_df.groupby('sector')['cost_score'].agg(['mean', 'count']).to_dict('index')
        
        return {
            'avg_cost_score': round(avg_cost, 2),
            'cost_distribution': cost_distribution,
            'potential_savings_percent': potential_savings,
            'baseline_cost': round(total_cost_baseline, 2),
            'optimized_cost': round(total_cost_optimized, 2),
            'savings_amount': round(total_cost_baseline - total_cost_optimized, 2),
            'sector_breakdown': {
                sector: {
                    'avg_cost': round(data['mean'], 2),
                    'product_count': int(data['count'])
                }
                for sector, data in sector_costs.items()
            }
        }
    
    def get_material_trends(self):
        """Calculate material usage trends"""
        material_stats = self.materials_df.groupby('material_type').agg({
            'material_id': 'count',
            'co2_emission_score': 'mean',
            'biodegradability_score': 'mean',
            'recyclability_percent': 'mean'
        }).round(2)
        
        material_stats.columns = ['count', 'avg_co2', 'avg_biodegradability', 'avg_recyclability']
        trend_data = material_stats.to_dict('index')
        
        top_materials = self.materials_df.nlargest(5, 'biodegradability_score')[
            ['material_name', 'material_type', 'biodegradability_score', 'recyclability_percent']
        ].to_dict('records')
        
        return {
            'material_types': trend_data,
            'total_materials': len(self.materials_df),
            'top_materials': top_materials
        }
    
    def get_sustainability_score(self):
        """Calculate overall sustainability score (0-100)"""
        co2_weight = 0.35
        biodegradability_weight = 0.35
        recyclability_weight = 0.30
        
        co2_norm = (100 - self.products_df['co2_emission_score'].mean())
        bio_norm = self.products_df['biodegradability_score'].mean()
        recycle_norm = self.products_df['recyclability_percent'].mean()
        
        score = (
            co2_norm * co2_weight +
            bio_norm * biodegradability_weight +
            recycle_norm * recyclability_weight
        )
        
        return round(score, 2)
    
    def create_co2_reduction_chart(self, output_file='co2_chart.html'):
        metrics = self.get_co2_metrics()
        baseline = metrics['total_emissions_baseline']
        optimized = metrics['total_emissions_optimized']
        
        fig = go.Figure(data=[
            go.Bar(name='Current', x=['Total CO2 Emissions'], y=[baseline], marker_color='#e74c3c'),
            go.Bar(name='Optimized', x=['Total CO2 Emissions'], y=[optimized], marker_color='#2ecc71')
        ])
        fig.update_layout(
            title='CO2 Emission Reduction Potential', barmode='group',
            yaxis_title='CO2 Equivalent (kg)', template='plotly_white', height=500
        )
        fig.write_html(output_file)
        return output_file
    
    def create_cost_savings_chart(self, output_file='cost_chart.html'):
        metrics = self.get_cost_metrics()
        fig = go.Figure(data=[go.Pie(
            labels=['Optimized Cost', 'Savings'],
            values=[metrics['optimized_cost'], metrics['savings_amount']],
            marker=dict(colors=['#3498db', '#27ae60']), hole=.3
        )])
        fig.update_layout(title=f"Cost Savings Potential: {metrics['potential_savings_percent']}%", height=500)
        fig.write_html(output_file)
        return output_file
    
    def create_material_trends_chart(self, output_file='material_trends.html'):
        data = self.get_material_trends()
        material_types = list(data['material_types'].keys())
        co2_scores = [data['material_types'][mt]['avg_co2'] for mt in material_types]
        biodeg_scores = [data['material_types'][mt]['avg_biodegradability'] for mt in material_types]
        recycle_scores = [data['material_types'][mt]['avg_recyclability'] for mt in material_types]
        
        fig = make_subplots(specs=[[{"secondary_y": False}]])
        fig.add_trace(go.Bar(x=material_types, y=co2_scores, name='Avg CO2 Score', marker_color='#e74c3c'))
        fig.add_trace(go.Scatter(x=material_types, y=biodeg_scores, name='Biodegradability %', line=dict(color='#27ae60', width=3), mode='lines+markers'))
        fig.add_trace(go.Scatter(x=material_types, y=recycle_scores, name='Recyclability %', line=dict(color='#3498db', width=3), mode='lines+markers'))
        fig.update_layout(title='Material Usage Trends', yaxis_title='Score (%)', height=500, template='plotly_white')
        fig.write_html(output_file)
        return output_file
    
    def create_sector_comparison_chart(self, output_file='sector_chart.html'):
        sector_data = self.products_df.groupby('sector').agg({
            'product_id': 'count', 'co2_emission_score': 'mean'
        }).round(2).reset_index()
        sector_data.columns = ['Sector', 'Product Count', 'Avg CO2']
        
        fig = make_subplots(rows=1, cols=2, specs=[[{"type": "bar"}, {"type": "scatter"}]])
        fig.add_trace(go.Bar(x=sector_data['Sector'], y=sector_data['Product Count'], name='Product Count', marker_color='#9b59b6'), row=1, col=1)
        fig.add_trace(go.Scatter(x=sector_data['Sector'], y=sector_data['Avg CO2'], name='Avg CO2 Score', mode='lines+markers', line=dict(color='#e74c3c', width=2)), row=1, col=2)
        fig.update_layout(title='Sector-wise Analysis', height=500, template='plotly_white')
        fig.write_html(output_file)
        return output_file
    
    def create_sustainability_gauge(self, output_file='sustainability_gauge.html'):
        score = self.get_sustainability_score()
        fig = go.Figure(go.Indicator(
            mode="gauge+number+delta", value=score,
            domain={'x': [0, 1], 'y': [0, 1]},
            title={'text': "Sustainability Score"}, delta={'reference': 50},
            gauge={
                'axis': {'range': [None, 100]}, 'bar': {'color': "#2ecc71"},
                'steps': [
                    {'range': [0, 25], 'color': "#e74c3c"},
                    {'range': [25, 50], 'color': "#f39c12"},
                    {'range': [50, 75], 'color': "#f1c40f"},
                    {'range': [75, 100], 'color': "#27ae60"}
                ]
            }
        ))
        fig.update_layout(height=500)
        fig.write_html(output_file)
        return output_file
    
    def get_all_metrics_json(self):
        return {
            'timestamp': datetime.now().isoformat(),
            'sustainability_score': self.get_sustainability_score(),
            'co2_metrics': self.get_co2_metrics(),
            'cost_metrics': self.get_cost_metrics(),
            'material_trends': self.get_material_trends(),
            'summary': {
                'total_products': len(self.products_df),
                'total_materials': len(self.materials_df),
                'data_loaded_from_file': self.data_loaded
            }
        }
    
    def export_to_excel(self, output_file='Sustainability_Report.xlsx'):
        """Export comprehensive report to Excel. Raises RuntimeError on failure."""
        # ✅ FIX: Raise exceptions instead of returning error strings
        if not HAS_OPENPYXL:
            raise RuntimeError(
                "openpyxl is not installed. Run: pip install openpyxl"
            )
        
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils.dataframe import dataframe_to_rows

        wb = Workbook()
        
        metrics = self.get_all_metrics_json()
        co2 = metrics['co2_metrics']
        cost = metrics['cost_metrics']

        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        # ── Summary Sheet ──────────────────────────────────────────────────────
        ws_summary = wb.active
        ws_summary.title = "Summary"
        ws_summary.append(['Metric', 'Value', 'Unit'])
        
        summary_data = [
            ['Sustainability Score',       metrics['sustainability_score'],              '%'],
            ['CO2 Reduction Potential',    co2['reduction_potential_percent'],           '%'],
            ['Current Total Emissions',    co2['total_emissions_baseline'],              'kg CO2'],
            ['Optimized Emissions',        co2['total_emissions_optimized'],             'kg CO2'],
            ['Cost Savings Potential',     cost['potential_savings_percent'],            '%'],
            ['Savings Amount',             f"Rs.{cost['savings_amount']:,.2f}",          'INR'],
            ['Total Products',             metrics['summary']['total_products'],         'units'],
            ['Total Materials',            metrics['summary']['total_materials'],        'units'],
        ]
        for row in summary_data:
            ws_summary.append(row)

        for cell in ws_summary[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        ws_summary.column_dimensions['A'].width = 30
        ws_summary.column_dimensions['B'].width = 20
        ws_summary.column_dimensions['C'].width = 15

        # ── Products Sheet ─────────────────────────────────────────────────────
        ws_products = wb.create_sheet("Products")
        for r_idx, row in enumerate(dataframe_to_rows(self.products_df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws_products.cell(row=r_idx, column=c_idx, value=value)
                if r_idx == 1:
                    cell.fill = header_fill
                    cell.font = header_font

        # ── Materials Sheet ────────────────────────────────────────────────────
        ws_materials = wb.create_sheet("Materials")
        for r_idx, row in enumerate(dataframe_to_rows(self.materials_df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws_materials.cell(row=r_idx, column=c_idx, value=value)
                if r_idx == 1:
                    cell.fill = header_fill
                    cell.font = header_font

        # ── CO2 Analysis Sheet ─────────────────────────────────────────────────
        ws_co2 = wb.create_sheet("CO2 Analysis")
        ws_co2.append(['CO2 Analysis Report'])
        ws_co2.append([])
        ws_co2.append(['Metric', 'Value'])
        ws_co2.append(['Average Product CO2 Score',  co2['avg_product_score']])
        ws_co2.append(['Average Material CO2 Score', co2['avg_material_score']])
        ws_co2.append(['Reduction Potential (%)',     co2['reduction_potential_percent']])
        ws_co2.append(['Current Baseline',            co2['total_emissions_baseline']])
        ws_co2.append(['Optimized Total',             co2['total_emissions_optimized']])

        # ── Cost Analysis Sheet ────────────────────────────────────────────────
        ws_cost = wb.create_sheet("Cost Analysis")
        ws_cost.append(['Cost Analysis Report'])
        ws_cost.append([])
        ws_cost.append(['Metric', 'Value'])
        ws_cost.append(['Average Cost Score',     cost['avg_cost_score']])
        ws_cost.append(['Potential Savings (%)',  cost['potential_savings_percent']])
        ws_cost.append(['Baseline Total Cost',    f"Rs.{cost['baseline_cost']:,.2f}"])
        ws_cost.append(['Optimized Total Cost',   f"Rs.{cost['optimized_cost']:,.2f}"])
        ws_cost.append(['Savings Amount',         f"Rs.{cost['savings_amount']:,.2f}"])

        wb.save(output_file)
        return f"Excel report exported to: {output_file}"
    
    def export_to_pdf(self, output_file='Sustainability_Report.pdf'):
        """Export comprehensive report to PDF. Raises RuntimeError on failure."""
        # ✅ FIX: Raise exceptions instead of returning error strings
        if not HAS_REPORTLAB:
            raise RuntimeError(
                "reportlab is not installed. Run: pip install reportlab"
            )
        
        doc = SimpleDocTemplate(
            output_file, pagesize=letter,
            rightMargin=0.5*inch, leftMargin=0.5*inch,
            topMargin=0.75*inch, bottomMargin=0.75*inch
        )
        
        elements = []
        styles = getSampleStyleSheet()
        
        elements.append(Paragraph("EcopackAI - Sustainability Report", styles['Title']))
        elements.append(Paragraph(
            f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", styles['Normal']
        ))
        elements.append(Spacer(1, 0.3*inch))
        
        metrics = self.get_all_metrics_json()
        co2  = metrics['co2_metrics']
        cost = metrics['cost_metrics']
        
        # ── Sustainability Score ───────────────────────────────────────────────
        elements.append(Paragraph("Sustainability Score", styles['Heading2']))
        status_text = 'Good' if metrics['sustainability_score'] >= 70 else 'Fair'
        score_table = Table(
            [['Overall Score', f"{metrics['sustainability_score']}/100"],
             ['Status', status_text]],
            colWidths=[2*inch, 2*inch]
        )
        score_table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#4472C4')),
            ('TEXTCOLOR',  (0,0), (-1,0), colors.whitesmoke),
            ('ALIGN',      (0,0), (-1,-1), 'CENTER'),
            ('FONTNAME',   (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE',   (0,0), (-1,0), 12),
            ('BOTTOMPADDING', (0,0), (-1,0), 12),
            ('BACKGROUND', (0,1), (-1,-1), colors.beige),
            ('GRID',       (0,0), (-1,-1), 1, colors.grey),
        ]))
        elements.append(score_table)
        elements.append(Spacer(1, 0.2*inch))
        
        # ── CO2 Metrics ───────────────────────────────────────────────────────
        elements.append(Paragraph("CO2 Reduction Metrics", styles['Heading2']))
        emission_reduction = co2['total_emissions_baseline'] - co2['total_emissions_optimized']
        co2_table = Table(
            [['Metric', 'Value'],
             ['Reduction Potential',    f"{co2['reduction_potential_percent']}%"],
             ['Current Total Emissions',f"{co2['total_emissions_baseline']} kg CO2"],
             ['Optimized Emissions',    f"{co2['total_emissions_optimized']} kg CO2"],
             ['Emission Reduction',     f"{emission_reduction:.2f} kg CO2"]],
            colWidths=[3*inch, 2*inch]
        )
        co2_table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#e74c3c')),
            ('TEXTCOLOR',  (0,0), (-1,0), colors.whitesmoke),
            ('ALIGN',      (0,0), (-1,-1), 'CENTER'),
            ('FONTNAME',   (0,0), (-1,0), 'Helvetica-Bold'),
            ('GRID',       (0,0), (-1,-1), 1, colors.grey),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.lightgrey, colors.white]),
        ]))
        elements.append(co2_table)
        elements.append(Spacer(1, 0.2*inch))
        
        # ── Cost Metrics ──────────────────────────────────────────────────────
        elements.append(Paragraph("Cost Savings Metrics", styles['Heading2']))
        cost_table = Table(
            [['Metric', 'Value'],
             ['Cost Savings Potential', f"{cost['potential_savings_percent']}%"],
             ['Baseline Cost',          f"Rs.{cost['baseline_cost']:,.2f}"],
             ['Optimized Cost',         f"Rs.{cost['optimized_cost']:,.2f}"],
             ['Total Savings',          f"Rs.{cost['savings_amount']:,.2f}"]],
            colWidths=[3*inch, 2*inch]
        )
        cost_table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#27ae60')),
            ('TEXTCOLOR',  (0,0), (-1,0), colors.whitesmoke),
            ('ALIGN',      (0,0), (-1,-1), 'CENTER'),
            ('FONTNAME',   (0,0), (-1,0), 'Helvetica-Bold'),
            ('GRID',       (0,0), (-1,-1), 1, colors.grey),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.lightgrey, colors.white]),
        ]))
        elements.append(cost_table)
        elements.append(Spacer(1, 0.3*inch))
        
        # ── Summary ───────────────────────────────────────────────────────────
        elements.append(Paragraph("Summary", styles['Heading2']))
        summary_text = (
            f"Based on the analysis of {metrics['summary']['total_products']} products and "
            f"{metrics['summary']['total_materials']} materials, the EcopackAI system recommends "
            f"implementing sustainable packaging solutions. This can result in:<br/><br/>"
            f"<b>CO2 Reduction:</b> {co2['reduction_potential_percent']}% reduction in emissions<br/>"
            f"<b>Cost Savings:</b> Rs.{cost['savings_amount']:,.2f} ({cost['potential_savings_percent']}%)<br/>"
            f"<b>Sustainability Score:</b> {metrics['sustainability_score']}/100"
        )
        elements.append(Paragraph(summary_text, styles['BodyText']))
        
        # ✅ FIX: Let exceptions propagate naturally — don't swallow them
        doc.build(elements)
        return f"PDF report exported to: {output_file}"


# ── Utility functions ──────────────────────────────────────────────────────────

def generate_dashboard_html(dashboard, output_dir='./'):
    os.makedirs(output_dir, exist_ok=True)
    return {
        'CO2 Reduction':    dashboard.create_co2_reduction_chart(os.path.join(output_dir, 'co2_reduction.html')),
        'Cost Savings':     dashboard.create_cost_savings_chart(os.path.join(output_dir, 'cost_savings.html')),
        'Material Trends':  dashboard.create_material_trends_chart(os.path.join(output_dir, 'material_trends.html')),
        'Sector Analysis':  dashboard.create_sector_comparison_chart(os.path.join(output_dir, 'sector_analysis.html')),
        'Sustainability Score': dashboard.create_sustainability_gauge(os.path.join(output_dir, 'sustainability_score.html')),
    }


def export_dashboard_reports(dashboard, output_dir='./'):
    os.makedirs(output_dir, exist_ok=True)
    return {
        'excel': dashboard.export_to_excel(os.path.join(output_dir, 'Sustainability_Report.xlsx')),
        'pdf':   dashboard.export_to_pdf(os.path.join(output_dir, 'Sustainability_Report.pdf')),
    }


if __name__ == "__main__":
    import json as _json

    print("Initializing EcopackAI Dashboard...")
    dashboard = EcopackDashboard()

    print("\n=== Dashboard Metrics ===")
    metrics = dashboard.get_all_metrics_json()

    def _to_serializable(o):
        if isinstance(o, (np.integer,)):  return int(o)
        if isinstance(o, (np.floating,)): return float(o)
        return str(o)

    print(_json.dumps(metrics, indent=2, default=_to_serializable))

    print("\nGenerating interactive charts...")
    charts = generate_dashboard_html(dashboard, output_dir='c:/EcopackAI/Dashboard')
    for name, file in charts.items():
        print(f"  {name}: {file}")

    print("\nExporting reports...")
    reports = export_dashboard_reports(dashboard, output_dir='c:/EcopackAI/Dashboard')
    for fmt, result in reports.items():
        print(f"  {fmt.upper()}: {result}")

    print("\nDashboard generation complete!")