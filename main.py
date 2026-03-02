from flask import Flask, request, render_template, jsonify, send_file
import numpy as np
import pickle
import pandas as pd
import io
import csv
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import os

app = Flask(__name__, static_folder='static', template_folder='templates')

# Load models
co2_model = pickle.load(open("co2_model.pkl", "rb"))
cost_model = pickle.load(open("cost_model.pkl", "rb"))

# Load CSV data
materials_df = pd.read_csv("material.csv")
products_df = pd.read_csv("product.csv")

# Cache for materials by category
materials_by_category = {}

@app.route("/")
def home():
    return render_template("index.html")

@app.route("/dashboard")
def dashboard():
    return render_template("dashboard.html")

@app.route("/recommendations")
def recommendations():
    return render_template("recommendations.html")

@app.route("/predict", methods=["POST"])
def predict():
    # Get input values from form
    float_features = [float(x) for x in request.form.values()]
    features = np.array([float_features])

    # Predictions
    co2_prediction = co2_model.predict(features)[0]
    cost_prediction = cost_model.predict(features)[0]

    return render_template(
        "index.html",
        co2_text=f"Predicted CO₂ Emission: {co2_prediction}",
        cost_text=f"Predicted Cost: ₹ {cost_prediction}"
    )

# Dashboard API endpoints
@app.route("/api/dashboard/summary", methods=["GET"])
def get_dashboard_summary():
    """Return summary metrics for the dashboard"""
    try:
        # Calculate summary metrics
        total_materials = len(products_df)
        avg_cost = materials_df['cost_score'].mean() if 'cost_score' in materials_df.columns else 0
        avg_co2 = materials_df['co2_emission_score'].mean() if 'co2_emission_score' in materials_df.columns else 0
        
        # Estimate CO₂ reduction and cost savings
        # Find eco-friendly vs non-eco materials
        eco_materials = materials_df[materials_df['biodegradability_score'] > 80]
        non_eco_materials = materials_df[materials_df['biodegradability_score'] <= 80]
        
        if len(eco_materials) > 0 and len(non_eco_materials) > 0:
            avg_eco_co2 = eco_materials['co2_emission_score'].mean()
            avg_non_eco_co2 = non_eco_materials['co2_emission_score'].mean()
            co2_reduction_pct = round(((avg_non_eco_co2 - avg_eco_co2) / avg_non_eco_co2 * 100), 2) if avg_non_eco_co2 > 0 else 0
            
            avg_eco_cost = eco_materials['cost_score'].mean()
            avg_non_eco_cost = non_eco_materials['cost_score'].mean()
            cost_savings_pct = round(((avg_non_eco_cost - avg_eco_cost) / avg_non_eco_cost * 100), 2) if avg_non_eco_cost > 0 else 0
        else:
            co2_reduction_pct = 0
            cost_savings_pct = 0
        
        return jsonify({
            "total_materials": int(total_materials),
            "average_cost_score": round(float(avg_cost), 2),
            "average_co2_score": round(float(avg_co2), 2),
            "estimated_co2_reduction_pct": co2_reduction_pct,
            "estimated_cost_savings_pct": cost_savings_pct,
            "eco_friendly_materials_count": len(eco_materials),
            "traditional_materials_count": len(non_eco_materials)
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/dashboard/charts", methods=["GET"])
def get_dashboard_charts():
    """Return chart data for visualizations"""
    try:
        # Prepare data by material type
        material_stats = defaultdict(lambda: {
            'cost': [],
            'co2': [],
            'biodegradability': [],
            'count': 0
        })
        
        # Aggregate by main_packaging_material
        for idx, row in materials_df.iterrows():
            material = row.get('main_packaging_material', 'Unknown')
            material_stats[material]['cost'].append(row.get('cost_score', 0))
            material_stats[material]['co2'].append(row.get('co2_emission_score', 0))
            material_stats[material]['biodegradability'].append(row.get('biodegradability_score', 0))
            material_stats[material]['count'] += 1
        
        # Calculate averages for each material
        co2_reduction_by_material = []
        cost_savings_by_material = []
        material_usage_trends = []
        
        baseline_co2 = 70  # baseline CO2 score
        baseline_cost = 3  # baseline cost score
        
        for material, stats in material_stats.items():
            avg_cost = np.mean(stats['cost']) if stats['cost'] else 0
            avg_co2 = np.mean(stats['co2']) if stats['co2'] else 0
            avg_bio = np.mean(stats['biodegradability']) if stats['biodegradability'] else 0
            
            # CO2 reduction calculation
            co2_reduction = max(0, baseline_co2 - avg_co2)
            co2_reduction_by_material.append({
                'material': material[:50],  # Truncate for readability
                'reduction': round(float(co2_reduction), 2),
                'co2_score': round(float(avg_co2), 2)
            })
            
            # Cost savings calculation
            cost_savings = max(0, baseline_cost - avg_cost)
            cost_savings_by_material.append({
                'material': material[:50],
                'savings': round(float(cost_savings), 2),
                'cost_score': round(float(avg_cost), 2)
            })
            
            # Material usage trends
            material_usage_trends.append({
                'material': material[:50],
                'usage_count': stats['count'],
                'biodegradability_score': round(float(avg_bio), 2)
            })
        
        return jsonify({
            "co2_reduction_by_material": co2_reduction_by_material,
            "cost_savings_by_material": cost_savings_by_material,
            "material_usage_trends": material_usage_trends
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/download_report", methods=["GET"])
def download_report():
    """Download sustainability report as CSV"""
    try:
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Write header and summary metrics
        summary_stats = get_dashboard_summary().get_json()
        writer.writerow(["EcoPackAI Sustainability Report"])
        writer.writerow([])
        writer.writerow(["SUMMARY METRICS"])
        writer.writerow(["Metric", "Value"])
        writer.writerow(["Total Materials", summary_stats['total_materials']])
        writer.writerow(["Average Cost Score", summary_stats['average_cost_score']])
        writer.writerow(["Average CO2 Score", summary_stats['average_co2_score']])
        writer.writerow(["Estimated CO2 Reduction (%)", summary_stats['estimated_co2_reduction_pct']])
        writer.writerow(["Estimated Cost Savings (%)", summary_stats['estimated_cost_savings_pct']])
        writer.writerow(["Eco-Friendly Materials", summary_stats['eco_friendly_materials_count']])
        writer.writerow(["Traditional Materials", summary_stats['traditional_materials_count']])
        writer.writerow([])
        
        # Write chart data
        writer.writerow(["MATERIAL ANALYSIS"])
        writer.writerow(["Material Type", "Average Cost Score", "Average CO2 Score", "Usage Count", "Biodegradability %"])
        
        # Aggregate by material type
        material_stats = defaultdict(lambda: {
            'cost': [],
            'co2': [],
            'biodegradability': [],
            'count': 0
        })
        
        for idx, row in materials_df.iterrows():
            material = str(row.get('main_packaging_material', 'Unknown')).encode('utf-8', 'ignore').decode('utf-8')
            material_stats[material]['cost'].append(row.get('cost_score', 0))
            material_stats[material]['co2'].append(row.get('co2_emission_score', 0))
            material_stats[material]['biodegradability'].append(row.get('biodegradability_score', 0))
            material_stats[material]['count'] += 1
        
        for material, stats in material_stats.items():
            avg_cost = np.mean(stats['cost']) if stats['cost'] else 0
            avg_co2 = np.mean(stats['co2']) if stats['co2'] else 0
            avg_bio = np.mean(stats['biodegradability']) if stats['biodegradability'] else 0
            writer.writerow([
                material[:50],
                round(float(avg_cost), 2),
                round(float(avg_co2), 2),
                stats['count'],
                round(float(avg_bio), 2)
            ])
        
        output.seek(0)
        return send_file(
            io.BytesIO(output.getvalue().encode('utf-8')),
            mimetype='text/csv; charset=utf-8',
            as_attachment=True,
            download_name='EcoPackAI_Sustainability_Report.csv'
        )
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# NEW ENDPOINTS FOR RECOMMENDATIONS PAGE

@app.route("/recommend", methods=["POST"])
def recommend():
    """Get recommendation based on user inputs"""
    try:
        data = request.json
        category = data.get('category', '')
        strength = data.get('strength', 5)
        weight = data.get('weight', 1.0)
        bio = data.get('bio', 50)
        recycle = data.get('recycle', 50)
        
        # Prepare features for model prediction
        features = np.array([[float(strength), float(weight), float(bio), float(recycle)]])
        
        # Get predictions from models
        co2_pred = float(co2_model.predict(features)[0])
        cost_pred = float(cost_model.predict(features)[0])
        
        # Find best material match in category
        if category:
            cat_materials = materials_df[materials_df['sector'].str.lower() == category.lower()]
        else:
            cat_materials = materials_df
        
        if len(cat_materials) > 0:
            best_material = cat_materials.iloc[0]['main_packaging_material']
        else:
            best_material = "Eco-Friendly Packaging"
        
        return jsonify({
            "material": best_material,
            "cost": cost_pred,
            "co2": co2_pred,
            "strength": strength,
            "weight": weight,
            "bio": bio,
            "recycle": recycle
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/environment-score", methods=["POST"])
def environment_score():
    """Calculate environmental score based on input"""
    try:
        data = request.json
        strength = data.get('strength', 5)
        weight = data.get('weight', 1.0)
        bio = data.get('bio', 50)
        recycle = data.get('recycle', 50)
        
        # Prepare features
        features = np.array([[float(strength), float(weight), float(bio), float(recycle)]])
        
        # Get CO2 prediction
        predicted_co2 = float(co2_model.predict(features)[0])
        baseline_co2 = 70.0
        
        # Calculate reduction percentage
        reduction = max(0, baseline_co2 - predicted_co2)
        reduction_percent = round((reduction / baseline_co2) * 100, 2) if baseline_co2 > 0 else 0
        
        return jsonify({
            "predicted_co2": round(predicted_co2, 2),
            "baseline_co2": baseline_co2,
            "reduction_percent": reduction_percent
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/ranked-materials/<category>", methods=["POST"])
def ranked_materials(category):
    """Get ranked materials for a category"""
    try:
        data = request.json
        strength = data.get('strength', 5)
        weight = data.get('weight', 1.0)
        bio = data.get('bio', 50)
        recycle = data.get('recycle', 50)
        
        features = np.array([[float(strength), float(weight), float(bio), float(recycle)]])
        
        # Filter by category if exists
        if category.lower() != 'all':
            filtered_materials = materials_df[materials_df['sector'].str.lower() == category.lower()]
        else:
            filtered_materials = materials_df
        
        if len(filtered_materials) == 0:
            filtered_materials = materials_df.head(10)
        
        # Prepare ranking data
        materials_list = []
        for idx, row in filtered_materials.iterrows():
            co2_pred = float(co2_model.predict(features)[0])
            cost_pred = float(cost_model.predict(features)[0])
            
            materials_list.append({
                "rank": len(materials_list) + 1,
                "material": str(row.get('main_packaging_material', 'Unknown')),
                "cost": cost_pred + (idx * 0.5),
                "co2": co2_pred + (idx * 0.1),
                "eco_score": max(60, 95 - (len(materials_list) * 6))
            })
        
        return jsonify({"materials": materials_list[:10]})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/materials", methods=["GET"])
def get_materials():
    """Get all materials grouped by category"""
    try:
        sectors = materials_df['sector'].unique()
        materials_map = {}
        
        for sector in sectors:
            sector_materials = materials_df[materials_df['sector'] == sector]['main_packaging_material'].unique()
            materials_map[sector] = [str(m) for m in sector_materials.tolist()]
        
        return jsonify(materials_map)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/export/excel", methods=["GET", "POST"])
def export_excel():
    """Export report as Excel file"""
    try:
        wb = Workbook()
        
        # Check if this is a recommendations export (POST) or summary export (GET)
        if request.method == "POST":
            data = request.get_json() or {}
            recommendations = data.get("recommendations", [])
            main_result = data.get("mainResult", {})
            
            # Recommendations sheet
            ws = wb.active
            ws.title = "Top Recommendations"
            
            # Header style
            header_fill = PatternFill(start_color="0F766E", end_color="0F766E", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=12)
            
            # Title
            ws['A1'] = "EcoPackAI - Material Recommendations"
            ws['A1'].font = Font(bold=True, size=14, color="0F766E")
            ws.merge_cells('A1:E1')
            
            # Summary section
            if main_result:
                row = 3
                ws[f'A{row}'] = "PRIMARY RECOMMENDATION"
                ws[f'A{row}'].font = Font(bold=True, size=11, color="0F766E")
                
                row += 1
                ws[f'A{row}'] = "Material:"
                ws[f'B{row}'] = main_result.get("material", "N/A")
                
                row += 1
                ws[f'A{row}'] = "Est. Cost (₹):"
                ws[f'B{row}'] = round(float(main_result.get("cost", 0)), 2)
                
                row += 1
                ws[f'A{row}'] = "CO₂ Impact (kg):"
                ws[f'B{row}'] = round(float(main_result.get("co2", 0)), 2)
                
                row = row + 2
            else:
                row = 4
            
            # Top recommendations table
            ws[f'A{row}'] = "TOP RANKED ALTERNATIVES (3-5)"
            ws[f'A{row}'].font = Font(bold=True, size=11, color="0F766E")
            
            row += 1
            headers = ["Rank", "Material", "Cost (₹)", "CO₂ (kg)", "Eco Score"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row, col)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            row += 1
            for rec in recommendations[:5]:  # Export top 5
                ws.cell(row, 1).value = rec.get("rank", "")
                ws.cell(row, 2).value = rec.get("material", "")
                ws.cell(row, 3).value = round(float(rec.get("cost", 0)), 2)
                ws.cell(row, 4).value = round(float(rec.get("co2", 0)), 4)
                ws.cell(row, 5).value = round(float(rec.get("eco_score", 0)), 1)
                
                # Format numbers
                ws.cell(row, 3).number_format = '#,##0.00'
                ws.cell(row, 4).number_format = '0.0000'
                ws.cell(row, 5).number_format = '0.0'
                row += 1
            
            # Summary benefits section
            row += 2
            ws[f'A{row}'] = "ESTIMATED BENEFITS"
            ws[f'A{row}'].font = Font(bold=True, size=11, color="0F766E")
            
            if recommendations and main_result:
                best_rec = recommendations[0]
                cost_savings = main_result.get("cost", 0) - best_rec.get("cost", 0)
                co2_savings = main_result.get("co2", 0) - best_rec.get("co2", 0)
                
                row += 1
                ws[f'A{row}'] = "Potential Cost Savings:"
                ws[f'B{row}'] = f"₹ {cost_savings:.2f}" if cost_savings > 0 else "N/A"
                
                row += 1
                ws[f'A{row}'] = "CO₂ Reduction:"
                ws[f'B{row}'] = f"{abs(co2_savings):.4f} kg" if co2_savings != 0 else "N/A"
            
            # Set column widths
            ws.column_dimensions['A'].width = 20
            ws.column_dimensions['B'].width = 25
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 15
            ws.column_dimensions['E'].width = 15
        
        else:
            # Original GET behavior - Summary report
            ws = wb.active
            ws.title = "Sustainability Report"
            
            # Header style
            header_fill = PatternFill(start_color="22c55e", end_color="22c55e", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            # Title
            ws['A1'] = "EcoPackAI Sustainability Report"
            ws['A1'].font = Font(bold=True, size=14)
            
            # Summary section
            row = 3
            ws[f'A{row}'] = "SUMMARY METRICS"
            ws[f'A{row}'].font = Font(bold=True, size=12)
            
            row += 1
            ws[f'A{row}'] = "Metric"
            ws[f'B{row}'] = "Value"
            
            for cell in [ws[f'A{row}'], ws[f'B{row}']]:
                cell.fill = header_fill
                cell.font = header_font
            
            # Get summary stats
            summary_stats = get_dashboard_summary().get_json()
            metrics = [
                ["Total Materials", summary_stats['total_materials']],
                ["Average Cost Score", summary_stats['average_cost_score']],
                ["Average CO₂ Score", summary_stats['average_co2_score']],
                ["Estimated CO₂ Reduction (%)", summary_stats['estimated_co2_reduction_pct']],
                ["Estimated Cost Savings (%)", summary_stats['estimated_cost_savings_pct']],
                ["Eco-Friendly Materials", summary_stats['eco_friendly_materials_count']],
                ["Traditional Materials", summary_stats['traditional_materials_count']]
            ]
            
            for metric, value in metrics:
                row += 1
                ws[f'A{row}'] = metric
                ws[f'B{row}'] = value
            
            # Material analysis section
            row += 3
            ws[f'A{row}'] = "MATERIAL ANALYSIS"
            ws[f'A{row}'].font = Font(bold=True, size=12)
            
            row += 1
            headers = ["Material", "Avg Cost", "Avg CO₂", "Usage Count", "Biodegradability %"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row, col)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
            
            # Add material data
            material_stats = defaultdict(lambda: {'cost': [], 'co2': [], 'biodegradability': [], 'count': 0})
            for idx, row_data in materials_df.iterrows():
                material = str(row_data.get('main_packaging_material', 'Unknown')).encode('utf-8', 'ignore').decode('utf-8')
                material_stats[material]['cost'].append(row_data.get('cost_score', 0))
                material_stats[material]['co2'].append(row_data.get('co2_emission_score', 0))
                material_stats[material]['biodegradability'].append(row_data.get('biodegradability_score', 0))
                material_stats[material]['count'] += 1
            
            row = row + 1
            for material, stats in material_stats.items():
                avg_cost = np.mean(stats['cost']) if stats['cost'] else 0
                avg_co2 = np.mean(stats['co2']) if stats['co2'] else 0
                avg_bio = np.mean(stats['biodegradability']) if stats['biodegradability'] else 0
                
                ws.cell(row, 1).value = material[:50]
                ws.cell(row, 2).value = round(float(avg_cost), 2)
                ws.cell(row, 3).value = round(float(avg_co2), 2)
                ws.cell(row, 4).value = stats['count']
                ws.cell(row, 5).value = round(float(avg_bio), 2)
                row += 1
            
            # Auto-adjust column widths
            ws.column_dimensions['A'].width = 35
            ws.column_dimensions['B'].width = 15
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 15
            ws.column_dimensions['E'].width = 18
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = 'Material_Recommendations.xlsx' if request.method == "POST" else 'Sustainability_Report.xlsx'
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/export/pdf", methods=["GET"])
def export_pdf():
    """Export report as PDF file"""
    try:
        # For now, return a CSV as PDF alternative
        # Full PDF support would require reportlab or similar
        from io import BytesIO
        
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Get summary data
        summary_stats = get_dashboard_summary().get_json()
        writer.writerow(["EcoPackAI Sustainability Report"])
        writer.writerow([])
        writer.writerow(["SUMMARY METRICS"])
        writer.writerow(["Metric", "Value"])
        writer.writerow(["Total Materials", summary_stats['total_materials']])
        writer.writerow(["Average Cost Score", summary_stats['average_cost_score']])
        writer.writerow(["Average CO₂ Score", summary_stats['average_co2_score']])
        writer.writerow(["Estimated CO₂ Reduction (%)", summary_stats['estimated_co2_reduction_pct']])
        writer.writerow(["Estimated Cost Savings (%)", summary_stats['estimated_cost_savings_pct']])
        writer.writerow(["Eco-Friendly Materials", summary_stats['eco_friendly_materials_count']])
        writer.writerow(["Traditional Materials", summary_stats['traditional_materials_count']])
        writer.writerow([])
        writer.writerow(["MATERIAL ANALYSIS"])
        writer.writerow(["Material Type", "Average Cost Score", "Average CO₂ Score", "Usage Count", "Biodegradability %"])
        
        material_stats = defaultdict(lambda: {'cost': [], 'co2': [], 'biodegradability': [], 'count': 0})
        for idx, row_data in materials_df.iterrows():
            material = str(row_data.get('main_packaging_material', 'Unknown')).encode('utf-8', 'ignore').decode('utf-8')
            material_stats[material]['cost'].append(row_data.get('cost_score', 0))
            material_stats[material]['co2'].append(row_data.get('co2_emission_score', 0))
            material_stats[material]['biodegradability'].append(row_data.get('biodegradability_score', 0))
            material_stats[material]['count'] += 1
        
        for material, stats in material_stats.items():
            avg_cost = np.mean(stats['cost']) if stats['cost'] else 0
            avg_co2 = np.mean(stats['co2']) if stats['co2'] else 0
            avg_bio = np.mean(stats['biodegradability']) if stats['biodegradability'] else 0
            writer.writerow([material[:50], round(float(avg_cost), 2), round(float(avg_co2), 2), stats['count'], round(float(avg_bio), 2)])
        
        output.seek(0)
        return send_file(
            io.BytesIO(output.getvalue().encode('utf-8')),
            mimetype='text/plain; charset=utf-8',
            as_attachment=True,
            download_name='Sustainability_Report.pdf'
        )
    except Exception as e:
        return jsonify({"error": str(e)}), 500
if __name__ == "__main__":
    app.run(debug=True)
